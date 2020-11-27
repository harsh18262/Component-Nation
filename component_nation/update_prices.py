#for scraping price from amazon
from bs4 import BeautifulSoup
import requests
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import *
import re

# def fetch_prices(r,c):
#     driver=init_selenium()
 
#     loc = ("Pre-Build Pc.xlsx")
 
#     wb=load_workbook(loc)
#     sheet1 = wb["Sheet1"]
#     sheet2 = wb["Sheet2"]
   

#     url=sheet1.cell(r,c).value

#     driver.get(url)

#     soup = BeautifulSoup(driver.page_source, "html.parser")

#     price_row = soup.find(id="priceblock_ourprice_row")
#     if(price_row==None):
#         price_row = soup.find(id="priceblock_dealprice_row")
#         price = price_row.find(class_="a-size-medium a-color-price priceBlockDealPriceString").get_text()
#     else:
#         price = price_row.find(class_="a-size-medium a-color-price priceBlockBuyingPriceString").get_text()


#     sheet2.cell(r,c).value=price
#     wb.save(loc)
#     driver.close()

def init_selenium():
    req=requests.get('http://pubproxy.com/api/proxy?format=txt&type=https&google=true&country=IND')
   # PROXY ='103.242.105.121:8080'
    CHROME_PATH = '/usr/bin/google-chrome-stable'
    CHROMEDRIVER_PATH = '/usr/bin/chromedriver'
    WINDOW_SIZE = "1920,1080"
    chrome_options = Options()
   # chrome_options.add_argument("--headless")
   # chrome_options.add_argument("--window-size=%s" % WINDOW_SIZE)
    #chrome_options.add_argument('--proxy-server=%s' % PROXY)
    chrome_options.binary_location = CHROME_PATH
    driver = webdriver.Chrome(executable_path=CHROMEDRIVER_PATH,chrome_options=chrome_options)
    return driver


def fetch_prices(data):
    
    urls=[]
    prices=[]
    for item in data:
        urls.append(item.url)
    print("abcdefgh")
    for url in urls:
        
        print(url)
        driver=init_selenium()
        driver.get(url)

        soup = BeautifulSoup(driver.page_source, "html.parser")

        price_row = soup.find(id="priceblock_ourprice_row")
        if(price_row==None):
            price_row = soup.find(id="priceblock_dealprice_row")
            prices.append(price_row.find(class_="a-size-medium a-color-price priceBlockDealPriceString").get_text())
        else:
            prices.append(price_row.find(class_="a-size-medium a-color-price priceBlockBuyingPriceString").get_text())



        driver.close()

    return prices

def total(data):
    total=0
    for item in data:
        total = total + int(re.sub("[^0-9]","",item.price))
    total="{:,}".format(total)
    return total
