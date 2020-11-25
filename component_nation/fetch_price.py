
#for scraping price from amazon
from openpyxl import *
loc = ("Pre-Build Pc.xlsx")
wb=load_workbook(loc)

def fetch_prices(r,c):
    sheet2 = wb["Sheet2"]
    return(sheet2.cell(r,c).value)

def fetch_urls(r,c):
    sheet1 = wb["Sheet1"]

    return(sheet1.cell(r,c).value)


def get_prices(r,c):
    return(fetch_prices(r,c))

def get_urls():
    for i in range (2,10):
        urls.append(fetch_urls(2,i))


